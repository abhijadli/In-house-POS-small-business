from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, update
from app.schema.product import ProductCreate
from app.models.product import Products
from app.exception.product_exception import (
    InvalidProductError,
    ProductImportError,
)
from typing import Any, cast
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook


async def get_product_by_id(db: AsyncSession, id: int) -> Products | None:
    result = await db.execute(
        select(Products).where(Products.id == id, Products.is_deleted.is_(False))
    )
    result = result.scalar_one_or_none()
    if result:
        return result


# List all available products
async def get_all_products(db: AsyncSession) -> list[Products]:
    result = await db.execute(
        select(Products)
        .where(Products.is_deleted.is_(False))
        .order_by(Products.name.asc())
    )
    result = result.scalars().all()
    return list(result)


# Add product to catalogue
async def add_product_to_catalogue(
    db: AsyncSession, product_details: ProductCreate
) -> Products:
    product: Products = Products(
        name=product_details.name,
        description=product_details.description,
        price=product_details.price,
        discount=product_details.discount,
        inventory=product_details.inventory,
    )
    db.add(product)
    await db.commit()
    await db.refresh(product)
    return product


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _decimal_value(value: Any, row_number: int, field: str) -> Decimal:
    value = _excel_value(value)
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise ProductImportError(
            f"Row {row_number}: {field} must be a valid number."
        )


def _inventory_value(value: Any, row_number: int) -> int:
    value = _excel_value(value)
    if isinstance(value, bool):
        raise ProductImportError(f"Row {row_number}: inventory must be a whole number.")
    try:
        inventory = int(value)
    except (TypeError, ValueError):
        raise ProductImportError(f"Row {row_number}: inventory must be a whole number.")
    if Decimal(str(value)) != inventory or inventory < 0:
        raise ProductImportError(
            f"Row {row_number}: inventory must be a whole number >= 0."
        )
    return inventory


def _parse_product_rows(file_content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(
            filename=BytesIO(file_content), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ProductImportError(
            "The uploaded file is not a readable .xlsx workbook."
        ) from exc

    worksheet = workbook.active
    if worksheet is None:
        raise ProductImportError("The workbook must contain at least one worksheet.")
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ProductImportError("The workbook must contain a header row.")

    normalized_headers = [
        str(header).strip().lower().replace(" ", "_") if header is not None else ""
        for header in headers
    ]
    required_headers = {"name", "price", "discount", "inventory"}
    if not required_headers.issubset(normalized_headers):
        missing = sorted(required_headers.difference(normalized_headers))
        raise ProductImportError(f"Missing required columns: {', '.join(missing)}.")

    column_indexes = {
        header: normalized_headers.index(header)
        for header in {"name", "description", "price", "discount", "inventory"}
        if header in normalized_headers
    }
    parsed_products: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        values = {}
        for field in ("name", "description", "price", "discount", "inventory"):
            index = column_indexes.get(field)
            values[field] = (
                _excel_value(row[index])
                if index is not None and index < len(row)
                else None
            )
        if all(value is None for value in values.values()):
            continue

        name = values["name"]
        if not isinstance(name, str):
            raise ProductImportError(f"Row {row_number}: name is required.")
        if len(name) > 255:
            raise ProductImportError(
                f"Row {row_number}: name must be at most 255 characters."
            )

        description = values["description"]
        if description is not None and not isinstance(description, str):
            raise ProductImportError(f"Row {row_number}: description must be text.")
        if description is not None and len(description) > 255:
            raise ProductImportError(
                f"Row {row_number}: description must be at most 255 characters."
            )

        price = _decimal_value(values["price"], row_number, "price")
        discount = _decimal_value(values["discount"], row_number, "discount")
        if price <= 0:
            raise ProductImportError(f"Row {row_number}: price must be greater than 0.")
        if discount < 0:
            raise ProductImportError(f"Row {row_number}: discount must be >= 0.")

        parsed_products.append(
            {
                "name": name,
                "description": description,
                "price": price,
                "discount": discount,
                "inventory": _inventory_value(values["inventory"], row_number),
            }
        )

    if not parsed_products:
        raise ProductImportError("The workbook does not contain any product rows.")
    return parsed_products


async def import_products_from_file(
    db: AsyncSession, content: bytes
) -> list[Products]:
    product_values = _parse_product_rows(content)
    products = [Products(**values) for values in product_values]
    try:
        db.add_all(products)
        await db.commit()
    except Exception:
        await db.rollback()
        raise
    for product in products:
        await db.refresh(product)
    return products


# Update a product inventory by its id
async def patch_inventory(db: AsyncSession, id: int, inventory: int) -> Products:
    product: Products | None = await get_product_by_id(db, id)
    if not product:
        raise InvalidProductError()
    setattr(product, "inventory", inventory)
    await db.commit()
    await db.refresh(product)
    return product


# Update a product inventory by its id
async def patch_product(db: AsyncSession, id: int, details: dict[str, Any]) -> Products:
    product: Products | None = await get_product_by_id(db, id)
    if not product:
        raise InvalidProductError()
    for field, values in details.items():
        setattr(product, field, values)
    await db.commit()
    await db.refresh(product)
    return product


# Remove product from catalogue
async def remove_product_from_catalogue(db: AsyncSession, id: int) -> Products:
    product = await get_product_by_id(db, id)
    if not product:
        raise InvalidProductError()
    setattr(product, "is_deleted", True)
    await db.commit()
    await db.refresh(product)
    return product


# Restock product after order failure
async def restock_product(db: AsyncSession, product_id: int, quantity: int) -> None:
    await db.execute(
        update(Products)
        .where(Products.id == product_id)
        .values(inventory=Products.inventory + quantity)
    )
