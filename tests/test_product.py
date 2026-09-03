import pytest
from io import BytesIO

from openpyxl import Workbook

from conftest import auth_header, create_product


def excel_file(
    *rows: tuple[object, ...],
    headers: tuple[str, ...] | None = None,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        headers or ("name", "description", "price", "discount", "inventory")
    )
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def upload_file(filename: str, content: bytes) -> dict:
    return {
        "file": (
            filename,
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }


@pytest.mark.asyncio
async def test_import_products_requires_super(client, user_token):
    response = await client.post(
        "/products/import_file",
        files=upload_file("products.xlsx", excel_file(("Widget", None, 10, 0, 1))),
        headers=auth_header(user_token),
    )

    assert response.status_code == 403


@pytest.mark.asyncio
async def test_import_products_rejects_non_xlsx_file(super_client, super_token):
    response = await super_client.post(
        "/products/import_file",
        files={
            "file": (
                "products.csv",
                b"name,price,discount,inventory\nWidget,10,0,1",
                "text/csv",
            )
        },
        headers=auth_header(super_token),
    )

    assert response.status_code == 415
    assert response.json()["detail"] == "Only excel file supported."


@pytest.mark.asyncio
async def test_import_products_rejects_corrupt_file(super_client, super_token):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file("products.xlsx", b"not-a-valid-xlsx"),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert "not a readable .xlsx workbook" in response.json()["detail"]


@pytest.mark.asyncio
async def test_import_products_rejects_missing_required_columns(
    super_client, super_token
):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file(("Widget", None, 10, 0, 1), headers=("name", "description")),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert "Missing required columns" in detail
    assert "price" in detail
    assert "discount" in detail
    assert "inventory" in detail


@pytest.mark.asyncio
async def test_import_products_rejects_workbook_without_data_rows(
    super_client, super_token
):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file("products.xlsx", excel_file()),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert response.json()["detail"] == "The workbook does not contain any product rows."


@pytest.mark.asyncio
async def test_import_products_skips_blank_rows(super_client, super_token):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(("name", "description", "price", "discount", "inventory"))
    worksheet.append(("First", None, 10, 0, 1))
    worksheet.append((None, None, None, None, None))
    worksheet.append(("Second", "desc", 20, 1, 2))
    output = BytesIO()
    workbook.save(output)

    response = await super_client.post(
        "/products/import_file",
        files=upload_file("products.xlsx", output.getvalue()),
        headers=auth_header(super_token),
    )

    assert response.status_code == 201, response.text
    body = response.json()
    assert [product["name"] for product in body] == ["First", "Second"]
    assert body[0]["description"] is None
    assert body[1]["description"] == "desc"
    assert float(body[0]["price"]) == 10.0
    assert float(body[1]["discount"]) == 1.0


@pytest.mark.asyncio
async def test_update_inventory_not_found(super_client, super_token):
    resp = await super_client.patch(
        "/products/9999/inventory",
        params={"new_inventory": 5},
        headers=auth_header(super_token),
    )
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_update_product_details_not_found(super_client, super_token):
    resp = await super_client.patch(
        "/products/9999/details",
        json={"name": "Missing"},
        headers=auth_header(super_token),
    )
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_import_products_rejects_missing_name(super_client, super_token):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file((None, None, 10, 0, 1)),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert "name is required" in response.json()["detail"]


@pytest.mark.asyncio
async def test_import_products_rejects_invalid_price_value(super_client, super_token):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file(("Widget", None, "not-a-number", 0, 1)),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert "price must be a valid number" in response.json()["detail"]


@pytest.mark.asyncio
async def test_import_products_rejects_negative_discount(super_client, super_token):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file(("Widget", None, 10, -1, 1)),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert "discount must be >= 0" in response.json()["detail"]


@pytest.mark.asyncio
async def test_import_products_rejects_non_integer_inventory(
    super_client, super_token
):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file(("Widget", None, 10, 0, 1.5)),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert "inventory must be a whole number" in response.json()["detail"]


@pytest.mark.asyncio
async def test_import_products_rejects_name_too_long(super_client, super_token):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file(("x" * 256, None, 10, 0, 1)),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert "name must be at most 255 characters" in response.json()["detail"]


@pytest.mark.asyncio
async def test_import_products_rejects_header_only_workbook(super_client, super_token):
    workbook = Workbook()
    workbook.active.append(
        ("name", "description", "price", "discount", "inventory")
    )
    output = BytesIO()
    workbook.save(output)

    response = await super_client.post(
        "/products/import_file",
        files=upload_file("products.xlsx", output.getvalue()),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert response.json()["detail"] == "The workbook does not contain any product rows."


@pytest.mark.asyncio
async def test_create_product_requires_super(client, user_token):
    resp = await client.post(
        "/products",
        json={"name": "X", "price": 10.0, "discount": 0.0, "inventory": 1},
        headers=auth_header(user_token),
    )
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_create_product(super_client, super_token):
    prod = await create_product(super_client, super_token, name="Widget")
    assert prod["name"] == "Widget"
    assert prod["inventory"] == 5
    assert prod["id"]


@pytest.mark.asyncio
async def test_import_products_from_excel(super_client, super_token, user_client):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file(
                ("Keyboard", "USB keyboard", 1200, 100, 8),
                ("Mouse", None, 600, 0, 15),
            ),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 201, response.text
    assert [product["name"] for product in response.json()] == ["Keyboard", "Mouse"]
    listed = await user_client.get("/products")
    assert len(listed.json()) == 2


@pytest.mark.asyncio
async def test_import_products_rejects_invalid_row_without_partial_write(
    super_client, super_token, user_client
):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file(("Valid", None, 10, 0, 1), ("Invalid", None, -1, 0, 1)),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    listed = await user_client.get("/products")
    assert listed.json() == []


@pytest.mark.asyncio
async def test_import_products_rejects_invalid_inventory(
    super_client, super_token, user_client
):
    response = await super_client.post(
        "/products/import_file",
        files=upload_file(
            "products.xlsx",
            excel_file(("Widget", None, 10, 0, -1)),
        ),
        headers=auth_header(super_token),
    )

    assert response.status_code == 422
    assert "inventory must be a whole number >= 0" in response.json()["detail"]
    listed = await user_client.get("/products")
    assert listed.json() == []


@pytest.mark.asyncio
async def test_create_product_invalid_price(super_client, super_token):
    resp = await super_client.post(
        "/products",
        json={"name": "X", "price": -1.0, "discount": 0.0, "inventory": 1},
        headers=auth_header(super_token),
    )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_list_products(super_client, super_token, user_client):
    await create_product(super_client, super_token, name="A")
    await create_product(super_client, super_token, name="B")
    resp = await user_client.get("/products")
    assert resp.status_code == 200
    names = [p["name"] for p in resp.json()]
    assert names == ["A", "B"]  # ordered by name asc


@pytest.mark.asyncio
async def test_get_product_by_id(super_client, super_token, user_client):
    prod = await create_product(super_client, super_token, name="Widget")
    resp = await user_client.get(f"/products/{prod['id']}")
    assert resp.status_code == 200
    assert resp.json()["name"] == "Widget"


@pytest.mark.asyncio
async def test_get_product_not_found(user_client):
    resp = await user_client.get("/products/9999")
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_update_inventory(super_client, super_token):
    prod = await create_product(super_client, super_token, inventory=5)
    resp = await super_client.patch(
        f"/products/{prod['id']}/inventory",
        params={"new_inventory": 20},
        headers=auth_header(super_token),
    )
    assert resp.status_code == 200
    assert resp.json()["inventory"] == 20


@pytest.mark.asyncio
async def test_update_product_details(super_client, super_token):
    prod = await create_product(super_client, super_token, name="Widget", price=100.0)
    resp = await super_client.patch(
        f"/products/{prod['id']}/details",
        json={"name": "Gadget", "price": 120.0},
        headers=auth_header(super_token),
    )
    assert resp.status_code == 201
    body = resp.json()
    assert body["name"] == "Gadget"
    assert float(body["price"]) == 120.0


@pytest.mark.asyncio
async def test_delete_product(super_client, super_token, user_client):
    prod = await create_product(super_client, super_token, name="Widget")
    resp = await super_client.delete(
        f"/products/{prod['id']}", headers=auth_header(super_token)
    )
    assert resp.status_code == 200
    # Soft-deleted products are hidden from listings/lookups.
    after = await user_client.get(f"/products/{prod['id']}")
    assert after.status_code == 404


@pytest.mark.asyncio
async def test_delete_product_not_found(super_client, super_token):
    resp = await super_client.delete(
        "/products/9999", headers=auth_header(super_token)
    )
    assert resp.status_code == 404
